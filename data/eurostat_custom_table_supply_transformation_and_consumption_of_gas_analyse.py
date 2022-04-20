import openpyxl

path = "."

wbs = openpyxl.load_workbook(filename=f"{path}/custom-table-supply-transformation-and-consumption-of-gas.xlsx", data_only=True,
                                read_only=True)
wbi = openpyxl.load_workbook(filename=f"{path}/custom-table-imports-of-natural-gas-by-partner-country.xlsx", data_only=True,
                                read_only=True)

sfile = open("sfile", "w")
dfile = open("debug", "w")




def sheetrange(begin, end):
    l = []
    for i in range(begin, end):
        l.append(f"Sheet {i}")
    return l

energy_categories = {"household": ['Sheet 20', 'Sheet 98', 'Sheet 99'],
                     "commercial": ['Sheet 96', 'Sheet 97'],
                     "electricity": ['Sheet 18', 'Sheet 19'],
                     "industry": sheetrange(21, 33) + ['Sheet 34', 'Sheet 55', 'Sheet 59'] + sheetrange(91, 92) + sheetrange(100, 103),
                     "other": sheetrange(49, 51) + ['Sheet 57'] + sheetrange(104, 106),
                     "transport": ['Sheet 56'] + sheetrange(87, 90) + sheetrange(93, 94) }

table_item_names = ["gas_imports_russia", "household_old", "household_new", "commercial_old", "commercial_new",
              "electricity_old", "electricity_new", "industry_old", "industry_new", "substitution", "balance", "transport_old", "transport_new", "other_old", "other_new"]
table_item_summary = ["household_old", "commercial_old", "electricity_old", "industry_old", "transport_old", "other_old"]



def summary(entry,country,ti):
    sum = 0.0
    print(f'{entry}\t', end = "", file=sfile)
    for t in table_item_summary:
        sum += ti[t]
    if sum == 0.0:
        return
    for t in table_item_summary:
        print(f'{round(ti[t])}\t', end="", file=sfile)
    print(f'{round(sum)}\t', end="", file=sfile)
    if isinstance(wbs["Sheet 15"][f"k{country}"].value, float):
        print(f'{round(wbs["Sheet 15"][f"k{country}"].value/1000)}\t', end="", file=sfile)
        print(f'{round(wbs["Sheet 16"][f"k{country}"].value/1000)}\t', end="", file=sfile)
    print("", file=sfile, flush=True)
    print(f'{entry}\t', end="", file=sfile)
    for t in table_item_summary:
        print(f'{round(ti[t]/sum*100)}\t', end="", file=sfile)
    print("", file=sfile, flush=True)


sums = {}

def scenario(savings_rates, year):
    print(f"\n\nCalculation: year {year}; savings rates: ", end="")
    for sr in savings_rates:
        print(f"{sr}: {round(savings_rates[sr]*100)} %, ", end="")
    print("\n")
    print("country\t", end="")
    for t in table_item_names:
        print(f'{t}\t', end = "")
        sums[t] = 0.0
    print("")

    for country in range(15, 55):
        country_name = ""
        if country == 49 or country == 50: # skip Serbia and Turkey
            continue
        ti = {}
        for t in table_item_names:
            ti[t] = 0.0
        unit_conversion = 0.001 # TJ to PJ

        colname = "k"
        if year == 2018:
            colname = "i"
        if year == 2019:
            colname = "j"
        if year == 2020:
            colname = "k"

        ti["gas_imports_russia"] = 0.0
        for s in ["Sheet 1", "Sheet 3"]:
            if wbi[s][f"k{country}"].value is not None and not isinstance(wbi[s][f"k{country}"].value, str):
                ti["gas_imports_russia"] += wbi[s][f"{colname}{country}"].value * unit_conversion
        sums["gas_imports_russia"] += ti["gas_imports_russia"]

        for ws in wbs.worksheets:
            if ws.title == "Sheet 1":
                country_name = ws[f"a{country}"].value
                if country_name == "Germany (until 1990 former territory of the FRG)":
                    country_name = "Germany"
                if country_name == "Kosovo (under United Nations Security Council Resolution 1244/99)":
                    country_name = "Kosovo"
                if country_name == "United Kingdom":
                    if year == 2020:
                        colname = "j" # 2020 data not yet available, use 2019 data instead
                print(f"{country_name}\t", end="")
            for energy_category in energy_categories:
                if ws[f"{colname}{country}"].value is not None and not isinstance(ws[f"{colname}{country}"].value, str) and ws.title in energy_categories[energy_category]:
                    tj = ws[f"{colname}{country}"].value  # value in TJ
                    pj = tj * unit_conversion # value in PJ
                    ti[f"{energy_category}_old"] += pj
                    sums[f"{energy_category}_old"] += pj
                    print(f'{round(pj)}\t{country_name} {energy_category}: {ws["c6"].value} year {year}', file=dfile, flush=True)
                    ti[f"{energy_category}_new"]+= pj * (1 - savings_rates[energy_category])
                    sums[f"{energy_category}_new"] += pj * (1 - savings_rates[energy_category])
                    ti["substitution"] += pj * savings_rates[energy_category]
                    sums["substitution"] += pj * savings_rates[energy_category]

        ti["balance"] = ti["substitution"] - ti["gas_imports_russia"]
        sums["balance"] += ti["balance"]

        for t in table_item_names:
            print(f'{round(ti[t])}\t', end="")
        print("")

        if (country == 19 or True):  # Germany
            summary(country_name, country, ti)

        if (country == 41): # After Sweden
            print("SUM EU\t", end="")
            for ti in table_item_names:
                print(f'{round(sums[ti])}\t', end="")
            print("")
            summary(country_name, country, sums)

        if (country == 54): # After Ukraine
            print("SUM Europe\t", end="")
            for ti in table_item_names:
                print(f'{round(sums[ti])}\t', end="")
            print("")

scenario({"household": 0.85, "commercial": 0.85, "electricity": 0.5, "industry": 0.08, "transport": 0.0, "other": 0.0}, 2020)
scenario({"household": 0.85, "commercial": 0.85, "electricity": 0.5, "industry": 0.08, "transport": 0.0, "other": 0.0}, 2019)
scenario({"household": 0.73, "commercial": 0.73, "electricity": 0.2, "industry": 0.08, "transport": 0.0, "other": 0.0}, 2020)
scenario({"household": 0.73, "commercial": 0.73, "electricity": 0.2, "industry": 0.08, "transport": 0.0, "other": 0.0}, 2019)


