import openpyxl
import time
import re

path = "."

wb = openpyxl.load_workbook(filename=f"{path}/custom-table-imports-of-natural-gas-by-partner-country.xlsx", data_only=True,
                                read_only=True)
f = open(f"{path}/output.tsv", "w")

for ws in wb.worksheets:
    if ws["k19"].value is not None:
        numval = ws["k19"].value # Germany
        print(f'{ws.title}\t{ws["c6"].value}: {ws["c7"].value}\t{numval}\t{ws["k12"].value}', file=f)
        print(f'{ws.title}\t{ws["c6"].value}: {ws["c7"].value}\t{numval}\t{ws["k12"].value}') # Europe