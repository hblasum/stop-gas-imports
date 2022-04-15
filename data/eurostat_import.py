import openpyxl

path = "."

wbs = openpyxl.load_workbook(filename=f"{path}/custom-table-supply-transformation-and-consumption-of-gas.xlsx", data_only=True,
                                read_only=True)
wbi = openpyxl.load_workbook(filename=f"{path}/custom-table-imports-of-natural-gas-by-partner-country.xlsx", data_only=True,
                                read_only=True)



energy_categories = {"household": ['Sheet 20', 'Sheet 96'], "commercial": ['Sheet 98'], "electricity": ['Sheet 18', 'Sheet 19'],
                     "industry": ['Sheet 21', 'Sheet 22', 'Sheet 23', 'Sheet 25', 'Sheet 35', 'Sheet 36', 'Sheet 37', 'Sheet 39', 'Sheet 42', 'Sheet 43', 'Sheet 45',
                                  'Sheet 48', 'Sheet 60', 'Sheet 62', 'Sheet 63', 'Sheet 64', 'Sheet 66', 'Sheet 68', 'Sheet 70', 'Sheet 71', 'Sheet 72', 'Sheet 74',
                                  'Sheet 76', 'Sheet 78', 'Sheet 80', 'Sheet 82', 'Sheet 91', 'Sheet 100', 'Sheet 102']}

table_item_names = ["gas_imports_russia", "household_old", "household_new", "commercial_old", "commercial_new",
              "electricity_old", "electricity_new", "industry_old", "industry_new", "substitution", "balance"]


sums = {}

def scenario(savings_rates):
    print(f"\n\nScenario: {savings_rates}")
    print("country\t", end="")
    for t in table_item_names:
        print(f'{t}\t', end = "")
        sums[t] = 0.0
    print("")

    for country in range(15, 55):
        if country == 49 or country == 50: # skip Serbia and Turkey
            continue
        ti = {}
        for t in table_item_names:
            ti[t] = 0.0
        unit_conversion = 0.001 # TJ to PJ

        ti["gas_imports_russia"] = 0.0
        for s in ["Sheet 1", "Sheet 3"]:
            if wbi[s][f"k{country}"].value is not None and not isinstance(wbi[s][f"k{country}"].value, str):
                ti["gas_imports_russia"] += wbi[s][f"k{country}"].value * unit_conversion
        sums["gas_imports_russia"] += ti["gas_imports_russia"]

        colname = "k"

        for ws in wbs.worksheets:
            if ws.title == "Sheet 1":
                country_name = ws[f"a{country}"].value
                if country_name == "Germany (until 1990 former territory of the FRG)":
                    country_name = "Germany"
                if country_name == "Kosovo (under United Nations Security Council Resolution 1244/99)":
                    country_name = "Kosovo"
                if country_name == "United Kingdom":
                    colname = "j" # 2020 data not yet available, use 2019 data instead
                print(f"{country_name}\t", end="")
            for energy_category in energy_categories:
                if ws[f"{colname}{country}"].value is not None and not isinstance(ws[f"{colname}{country}"].value, str) and ws.title in energy_categories[energy_category]:
                    tj = ws[f"{colname}{country}"].value  # value in TJ
                    pj = tj * unit_conversion # value in PJ
                    ti[f"{energy_category}_old"] += pj
                    sums[f"{energy_category}_old"] += pj
                    ti[f"{energy_category}_new"]+= pj * (1 - savings_rates[energy_category])
                    sums[f"{energy_category}_new"] += pj * (1 - savings_rates[energy_category])
                    ti["substitution"] += pj * savings_rates[energy_category]
                    sums["substitution"] += pj * savings_rates[energy_category]

        ti["balance"] = ti["substitution"] - ti["gas_imports_russia"]
        sums["balance"] += ti["balance"]

        for t in table_item_names:
            print(f'{round(ti[t])}\t', end="")
        print("")

        if (country == 40): # After Finland
            print("SUMs EU\t", end="")
            for ti in table_item_names:
                print(f'{round(sums[ti])}\t', end="")
            print("")

        if (country == 54): # After Ukraine
            print("SUMs Europe\t", end="")
            for ti in table_item_names:
                print(f'{round(sums[ti])}\t', end="")
            print("")

scenario({"household": 0.77, "commercial": 0.77, "electricity": 0.3, "industry": 0.2})
scenario({"household": 0.75, "commercial": 0.75, "electricity": 0.1, "industry": 0.1})